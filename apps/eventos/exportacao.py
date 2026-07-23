"""Geração de planilha (.xlsx) da lista de inscritos de um evento.

As colunas exportáveis são de dois tipos:
- **Base**: campos fixos da inscrição/usuário (`COLUNAS_BASE`).
- **Dinâmicas**: um por campo personalizado do evento (`CampoEvento`), com chave
  `campo_<id>`.

A view escolhe quais colunas incluir e em qual ordem (o usuário reordena por
drag-and-drop no front-end); aqui apenas resolvemos chave -> rótulo/valor.
"""
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _sim_nao(valor):
    return 'Sim' if valor else 'Não'


def _fmt_data_hora(dt):
    if not dt:
        return ''
    return timezone.localtime(dt).strftime('%d/%m/%Y %H:%M')


def _fmt_data(dt):
    if not dt:
        return ''
    return timezone.localtime(dt).strftime('%d/%m/%Y')


# (chave, rótulo, função(inscricao) -> valor)
COLUNAS_BASE = [
    ('nome_completo', 'Nome completo', lambda i: i.usuario.display_name),
    ('primeiro_nome', 'Primeiro nome', lambda i: i.usuario.first_name),
    ('sobrenome', 'Sobrenome', lambda i: i.usuario.last_name),
    ('usuario', 'Usuário', lambda i: i.usuario.username),
    ('email', 'E-mail', lambda i: i.usuario.email),
    ('whatsapp', 'WhatsApp', lambda i: i.usuario.whatsapp),
    ('tipo_usuario', 'Tipo de usuário', lambda i: i.usuario.get_tipo_display()),
    ('papel', 'Papel no evento', lambda i: i.get_papel_evento_display()),
    ('status', 'Status', lambda i: i.get_status_display()),
    ('pago', 'Pagamento confirmado', lambda i: _sim_nao(i.pago)),
    ('data_pagamento', 'Data do pagamento', lambda i: _fmt_data(i.data_pagamento)),
    ('credencial_validada', 'Credencial validada', lambda i: _sim_nao(i.credencial_validada)),
    ('credential_url', 'Link da credencial', lambda i: i.credential_url or ''),
    ('data_inscricao', 'Data da inscrição', lambda i: _fmt_data_hora(i.data_inscricao)),
    ('observacoes', 'Observações', lambda i: i.observacoes or ''),
    ('motivo_rejeicao', 'Motivo da rejeição', lambda i: i.motivo_rejeicao or ''),
]

# Colunas marcadas por padrão ao abrir o modal de exportação.
COLUNAS_PADRAO = ['nome_completo', 'email', 'whatsapp', 'papel', 'status']


def _getter_campo(campo_id):
    def getter(inscricao):
        for resposta in inscricao.respostas.all():
            if resposta.campo_id == campo_id:
                return resposta.valor
        return ''
    return getter


def colunas_disponiveis(evento):
    """Lista [(chave, rótulo)] de todas as colunas exportáveis do evento.

    Ordem: colunas base + um item por campo personalizado do evento.
    """
    colunas = [(chave, rotulo) for chave, rotulo, _ in COLUNAS_BASE]
    for campo in evento.campos_personalizados.all():
        colunas.append((f'campo_{campo.id}', campo.label))
    return colunas


def _resolver(evento):
    """Retorna (getters, rotulos) mapeando chave -> função e chave -> rótulo."""
    getters = {chave: fn for chave, _rotulo, fn in COLUNAS_BASE}
    rotulos = {chave: rotulo for chave, rotulo, _ in COLUNAS_BASE}
    for campo in evento.campos_personalizados.all():
        chave = f'campo_{campo.id}'
        getters[chave] = _getter_campo(campo.id)
        rotulos[chave] = campo.label
    return getters, rotulos


def gerar_planilha_bytes(evento, inscricoes, colunas):
    """Gera a planilha .xlsx e devolve os bytes prontos para download.

    `colunas` é a lista ordenada de chaves escolhidas pelo usuário. Chaves
    desconhecidas são ignoradas; se nenhuma for válida, usa `COLUNAS_PADRAO`.
    """
    getters, rotulos = _resolver(evento)
    colunas = [c for c in colunas if c in getters]
    if not colunas:
        colunas = [c for c in COLUNAS_PADRAO if c in getters]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inscritos'

    cabecalho = [rotulos[c] for c in colunas]
    ws.append(cabecalho)

    header_fill = PatternFill('solid', fgColor='1E3A8A')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center', horizontal='left')

    # Largura inicial baseada no cabeçalho; cresce conforme o conteúdo.
    larguras = [len(str(t)) for t in cabecalho]

    for inscricao in inscricoes:
        linha = []
        for idx, chave in enumerate(colunas):
            valor = getters[chave](inscricao)
            valor = '' if valor is None else str(valor)
            linha.append(valor)
            larguras[idx] = max(larguras[idx], len(valor))
        ws.append(linha)

    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(largura + 2, 10), 60)

    ws.freeze_panes = 'A2'

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
